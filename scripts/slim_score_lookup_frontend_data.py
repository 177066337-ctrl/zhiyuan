from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pdfplumber

from national_extraction_common import ROOT, dump_json, iso_now, load_json, normalize_text, safe_float, safe_int, write_markdown

APP_BASE = ROOT / "app" / "public" / "data" / "score-lookup"
APP_INDEX = APP_BASE / "index.json"
ADMISSIONS_DIR = APP_BASE / "admissions"
RANK_DIR = APP_BASE / "rank-tables"
ARCHIVE_DIR = ROOT / "data_work" / "frontend_score_lookup_archive"
ARCHIVE_ADMISSIONS = ARCHIVE_DIR / "admissions"
ARCHIVE_RANKS = ARCHIVE_DIR / "rank-tables"
AUDIT_JSON = ROOT / "data_work" / "enabled_score_lookup_audit.json"
SIZE_JSON = ROOT / "data_work" / "frontend_data_size_report.json"
SIZE_MD = ROOT / "docs" / "frontend_data_size_report.md"
ADMISSIONS_FOLDERS = [
    ROOT / "data_work" / "national_admissions_normalized",
    ROOT / "data_work" / "remaining_admissions_normalized",
]
WATERMARK_TOKEN_RE = re.compile(r"(^|\s)[江西省教育考试院](?=\s|$)")


def clean_pdf_text(value: Any) -> str:
    text = normalize_text(value).replace("\n", " ")
    text = WATERMARK_TOKEN_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def digits_only(value: Any) -> str:
    return "".join(ch for ch in clean_pdf_text(value) if ch.isdigit())


def total_json_size(folder: Path) -> int:
    return sum(path.stat().st_size for path in folder.rglob("*.json") if path.is_file())


def move_to_archive(path: Path, archive_root: Path) -> None:
    if not path.exists():
        return
    archive_root.mkdir(parents=True, exist_ok=True)
    target = archive_root / path.name
    if target.exists():
        target.unlink()
    path.replace(target)


def read_manifest_shards(path: Path) -> list[Path]:
    try:
        payload = load_json(path)
    except Exception:
        return []
    if isinstance(payload, dict) and isinstance(payload.get("shards"), list):
        return [APP_BASE.parent / shard for shard in payload["shards"]]
    return []


def find_dataset_file(dataset: dict[str, Any]) -> Path:
    for folder in ADMISSIONS_FOLDERS:
        for path in folder.glob("*.json"):
            records = load_json(path)
            if not isinstance(records, list) or not records:
                continue
            first = records[0]
            if (
                normalize_text(first.get("province")) == normalize_text(dataset["province"])
                and safe_int(first.get("year")) == safe_int(dataset["year"])
                and normalize_text(first.get("subject_type")) == normalize_text(dataset["subject_type"])
                and normalize_text(first.get("batch")) == normalize_text(dataset["batch"])
            ):
                return path
    raise FileNotFoundError(dataset["dataset_id"])


def corrected_record_from_source(
    record: dict[str, Any],
    excel_cache: dict[tuple[str, str, int], dict[str, Any]],
    pdf_cache: dict[tuple[str, int], tuple[list[str], list[list[Any]]]],
) -> dict[str, Any]:
    source_file = Path(record["source_file"])
    source_row = safe_int(record.get("source_row"))
    source_page = safe_int(record.get("source_page"))
    batch = normalize_text(record.get("batch"))

    if source_file.suffix.lower() in {".xlsx", ".xlsm"}:
        import openpyxl

        source_sheet = normalize_text(record.get("source_sheet")) or "Sheet1"
        cache_key = (str(source_file), source_sheet, source_row or 0)
        if cache_key not in excel_cache:
            workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
            worksheet = workbook[source_sheet if source_sheet in workbook.sheetnames else workbook.sheetnames[0]]
            headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            values = next(worksheet.iter_rows(min_row=source_row, max_row=source_row, values_only=True))
            excel_cache[cache_key] = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
        row = excel_cache[cache_key]
        return {
            "school_code": "",
            "school_name": normalize_text(row.get("学校")),
            "major_group_code": "",
            "major_group_name": "",
            "major_code": "",
            "major_name": normalize_text(row.get("专业")),
            "min_score": safe_float(row.get("最低分")),
            "min_rank": safe_int(row.get("最低位次")),
            "batch": normalize_text(row.get("批次")) or batch,
            "remarks": "",
            "source_file": normalize_text(record.get("source_file")),
            "confidence": normalize_text(record.get("confidence")) or "medium",
        }

    if source_file.suffix.lower() == ".pdf":
        cache_key = (str(source_file), source_page or 0)
        if cache_key not in pdf_cache:
            with pdfplumber.open(source_file) as pdf:
                page = pdf.pages[(source_page or 1) - 1]
                table = max(page.extract_tables() or [[]], key=len)
            headers = [clean_pdf_text(cell) for cell in table[0]]
            pdf_cache[cache_key] = (headers, table)
        headers, table = pdf_cache[cache_key]
        row = table[(source_row or 1) - 1]
        row_map = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        return {
            "school_code": digits_only(row_map.get("院校 代号")) or digits_only(row_map.get("院校代号")),
            "school_name": clean_pdf_text(row_map.get("院校名称")),
            "major_group_code": digits_only(row_map.get("专业组 代号")) or digits_only(row_map.get("专业组代号")),
            "major_group_name": clean_pdf_text(row_map.get("专业组 名称")) or clean_pdf_text(row_map.get("专业组名称")),
            "major_code": "",
            "major_name": "",
            "min_score": safe_float(clean_pdf_text(row_map.get("投档线"))),
            "min_rank": safe_int(clean_pdf_text(row_map.get("最低 投档排名")) or clean_pdf_text(row_map.get("最低投档排名"))),
            "batch": batch,
            "remarks": "",
            "source_file": normalize_text(record.get("source_file")),
            "confidence": normalize_text(record.get("confidence")) or "high",
        }

    raise ValueError(f"unsupported source type: {source_file}")


def rebuild_enabled_admissions(datasets: list[dict[str, Any]]) -> None:
    excel_cache: dict[tuple[str, str, int], dict[str, Any]] = {}
    pdf_cache: dict[tuple[str, int], tuple[list[str], list[list[Any]]]] = {}
    for dataset in datasets:
        source_path = find_dataset_file(dataset)
        records = load_json(source_path)
        compacted = [corrected_record_from_source(record, excel_cache, pdf_cache) for record in records]
        out_path = ADMISSIONS_DIR / Path(dataset["admissions_file"]).name
        out_path.write_text(json.dumps(compacted, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    index = load_json(APP_INDEX)
    audit = load_json(AUDIT_JSON)
    keep_ids = set(audit["release_keep_dataset_ids"])
    dataset_map = {item["dataset_id"]: item for item in index["datasets"]}
    keep_datasets = [dataset_map[dataset_id] for dataset_id in keep_ids if dataset_id in dataset_map]

    before_size = total_json_size(APP_BASE)

    ARCHIVE_ADMISSIONS.mkdir(parents=True, exist_ok=True)
    ARCHIVE_RANKS.mkdir(parents=True, exist_ok=True)

    keep_admission_names = {Path(item["admissions_file"]).name for item in keep_datasets}
    keep_rank_names = {Path(item["rank_table_file"]).name for item in keep_datasets if item.get("rank_table_file")}

    for path in list(ADMISSIONS_DIR.glob("*.json")):
        if path.name in keep_admission_names:
            continue
        for shard_path in read_manifest_shards(path):
            move_to_archive(shard_path, ARCHIVE_ADMISSIONS)
        move_to_archive(path, ARCHIVE_ADMISSIONS)

    for path in list(RANK_DIR.glob("*.json")):
        if path.name in keep_rank_names:
            continue
        move_to_archive(path, ARCHIVE_RANKS)

    rebuild_enabled_admissions(keep_datasets)

    for dataset in index["datasets"]:
        if dataset["dataset_id"] in keep_ids:
            dataset["enabled"] = True
            dataset["notes"] = normalize_text(dataset.get("notes"))
        else:
            dataset["enabled"] = False
            current_notes = normalize_text(dataset.get("notes"))
            dataset["notes"] = f"{current_notes}; release_hidden".strip("; ").strip()

    index["generated_at"] = iso_now()
    dump_json(APP_INDEX, index)

    after_size = total_json_size(APP_BASE)
    report = {
        "generated_at": iso_now(),
        "before_size_bytes": before_size,
        "after_size_bytes": after_size,
        "enabled_dataset_ids": sorted(keep_ids),
        "enabled_dataset_count": len(keep_ids),
        "admissions_files_remaining": sorted(path.name for path in ADMISSIONS_DIR.glob("*.json")),
        "rank_files_remaining": sorted(path.name for path in RANK_DIR.glob("*.json")),
    }
    dump_json(SIZE_JSON, report)

    lines = [
        "# Frontend Data Size Report",
        "",
        f"- 瘦身前体积：{before_size / 1024 / 1024:.2f} MB",
        f"- 瘦身后体积：{after_size / 1024 / 1024:.2f} MB",
        f"- 保留 enabled 数据集数量：{len(keep_ids)}",
        f"- 保留 admissions 文件数量：{len(report['admissions_files_remaining'])}",
        f"- 保留 rank_table 文件数量：{len(report['rank_files_remaining'])}",
        "",
        "## Kept Datasets",
        "",
    ]
    for dataset_id in sorted(keep_ids):
        lines.append(f"- {dataset_id}")
    lines.extend(
        [
            "",
            "## Remaining Frontend Files",
            "",
        ]
    )
    for name in report["admissions_files_remaining"]:
        lines.append(f"- admissions/{name}")
    for name in report["rank_files_remaining"]:
        lines.append(f"- rank-tables/{name}")
    write_markdown(SIZE_MD, "\n".join(lines))

    print(
        f"slimmed score-lookup data from {before_size / 1024 / 1024:.2f} MB "
        f"to {after_size / 1024 / 1024:.2f} MB; kept {len(keep_ids)} datasets"
    )


if __name__ == "__main__":
    main()
