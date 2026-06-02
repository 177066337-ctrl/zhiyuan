from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "app" / "public" / "data"

SCHOOL_SOURCE = ROOT / "全国普通高等学校名单.xls"
UNDERGRAD_MAJOR_SOURCE = ROOT / "普通高等学校本科专业目录.xlsx"
COLLEGE_MAJOR_SOURCE = ROOT / "普通高等学校高等职业教育（专科）专业目录.xlsx"


@dataclass
class ConversionResult:
    schools: list[dict[str, Any]]
    majors: list[dict[str, Any]]
    logs: list[str]


def stable_id(prefix: str, *parts: str) -> str:
    raw = "||".join((part or "").strip() for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\u3000", " ").replace("\n", " ").strip()
    return " ".join(text.split())


def write_json(path: Path, data: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_xls_rows(path: Path) -> list[tuple[int, list[str]]]:
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    rows: list[tuple[int, list[str]]] = []
    for idx in range(sheet.nrows):
        row = [clean_text(sheet.cell_value(idx, col)) for col in range(sheet.ncols)]
        rows.append((idx + 1, row))
    return rows


def load_xlsx_rows(path: Path, sheet_index: int = 0) -> tuple[str, list[tuple[int, list[str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[sheet_index]
    rows: list[tuple[int, list[str]]] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        rows.append((idx, [clean_text(cell) for cell in row]))
    return sheet.title, rows


def infer_school_level(value: str) -> str:
    if "本科" in value:
        return "本科"
    if "专科" in value:
        return "专科"
    return value or "其他"


def infer_ownership(department: str, note: str, school_name: str) -> str:
    merged = " ".join(part for part in [department, note, school_name] if part)
    if not merged:
        return "其他"
    if "中外合作" in merged:
        return "中外合作"
    if "港澳" in merged or "港澳台" in merged:
        return "内地与港澳台合作"
    if "民办" in merged:
        return "民办"
    if "独立学院" in merged:
        return "其他"
    public_markers = [
        "教育部",
        "教育厅",
        "教育局",
        "人民政府",
        "委员会",
        "中国科学院",
        "中国社会科学院",
        "国家",
        "省",
        "市",
        "自治区",
        "兵团",
    ]
    if any(marker in department for marker in public_markers):
        return "公办"
    return "其他"


def is_province_group(row: list[str]) -> bool:
    if not row:
        return False
    first = row[0]
    return bool(first) and "（" in first and "所）" in first and all(not cell for cell in row[1:])


def parse_province_group(value: str) -> str:
    return value.split("（", 1)[0].strip()


def convert_schools(logs: list[str]) -> list[dict[str, Any]]:
    rows = load_xls_rows(SCHOOL_SOURCE)
    schools: list[dict[str, Any]] = []
    current_province = ""
    header_seen = False

    for source_row, row in rows:
        padded = row + [""] * (6 - len(row))
        first = padded[0]

        if not any(padded):
            continue
        if first.startswith("全国普通高等学校名单"):
            logs.append(f"skip school title row {source_row}")
            continue
        if is_province_group(padded):
            current_province = parse_province_group(first)
            logs.append(f"province group row {source_row}: {current_province}")
            continue
        if padded[:6] == ["序号", "学校名称", "主管部门", "所在地", "办学层次", "备注"]:
            header_seen = True
            logs.append(f"school header row {source_row}")
            continue
        if not header_seen:
            continue

        school_name = padded[1]
        department = padded[2]
        location = padded[3]
        school_level = infer_school_level(padded[4])
        note = padded[5]

        if not school_name or not current_province:
            logs.append(f"skip school row {source_row}: missing school_name or province")
            continue

        city = location
        record = {
            "school_id": stable_id("school", school_name, current_province, city),
            "school_name": school_name,
            "province": current_province,
            "city": city,
            "department": department,
            "school_level": school_level,
            "school_type": "",
            "ownership": infer_ownership(department, note, school_name),
            "tags": [],
            "official_site": "",
            "source_file": SCHOOL_SOURCE.name,
            "source_row": source_row,
        }
        schools.append(record)

    logs.append(f"schools parsed: {len(schools)}")
    return schools


def convert_undergrad_majors(logs: list[str]) -> list[dict[str, Any]]:
    sheet_name, rows = load_xlsx_rows(UNDERGRAD_MAJOR_SOURCE)
    majors: list[dict[str, Any]] = []
    current_discipline = ""
    current_category = ""

    logs.append(f"undergrad sheet: {sheet_name}")

    for source_row, row in rows:
        padded = row + [""] * (2 - len(row))
        code = padded[0]
        name = padded[1]

        if not code and not name:
            continue
        if [code, name] == ["专业代码", "专业名称"]:
            logs.append(f"undergrad header row {source_row}")
            continue
        if not code or not name:
            logs.append(f"skip undergrad row {source_row}: incomplete row")
            continue

        normalized_code = code.upper()
        if len(normalized_code) == 2 and "学科门类" in name:
            current_discipline = name.replace("学科门类：", "").strip()
            current_category = ""
            continue
        if len(normalized_code) == 4:
            current_category = name
            continue
        if len(normalized_code) >= 6:
            majors.append(
                {
                    "major_id": stable_id("major", "本科", normalized_code, name),
                    "major_code": normalized_code,
                    "major_name": name,
                    "major_category": current_category,
                    "major_discipline": current_discipline,
                    "degree_level": "本科",
                    "degree": "",
                    "duration": "",
                    "subject_requirement": "",
                    "description": "",
                    "source_file": UNDERGRAD_MAJOR_SOURCE.name,
                    "source_row": source_row,
                }
            )
            continue

        logs.append(f"skip undergrad row {source_row}: unrecognized code {normalized_code}")

    logs.append(f"undergrad majors parsed: {len(majors)}")
    return majors


def convert_college_majors(logs: list[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(COLLEGE_MAJOR_SOURCE, read_only=True, data_only=True)
    majors: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    # Sheet 1: 主目录
    main_sheet = workbook.worksheets[0]
    current_discipline = ""
    current_category = ""
    header_seen = False
    for source_row, cells in enumerate(main_sheet.iter_rows(values_only=True), start=1):
        row = [clean_text(cell) for cell in cells]
        padded = row + [""] * (6 - len(row))
        if not any(padded):
            continue
        if padded[0] == "序号" and padded[1] == "专业代码":
            header_seen = True
            logs.append(f"college main header row {source_row}")
            continue
        if not header_seen:
            continue

        first = padded[0]
        code = padded[1]
        name = padded[2]
        old_code = padded[3]
        old_name = padded[4]
        adjust = padded[5]

        if first and "大类" in first and not code and not name:
            current_discipline = first
            current_category = ""
            continue
        if first and first[:4].isdigit() and "类" in first and not code and not name:
            current_category = first
            continue
        if not code and old_code and old_name:
            continue
        if not code or not name:
            logs.append(f"skip college main row {source_row}: missing code or name")
            continue

        major = {
            "major_id": stable_id("major", "专科", code, name),
            "major_code": code,
            "major_name": name,
            "major_category": current_category,
            "major_discipline": current_discipline,
            "degree_level": "专科",
            "degree": "",
            "duration": "",
            "subject_requirement": "",
            "description": "",
            "source_file": COLLEGE_MAJOR_SOURCE.name,
            "source_row": source_row,
        }
        if major["major_id"] not in seen_ids:
            majors.append(major)
            seen_ids.add(major["major_id"])

    # Sheet 2: 增补专业
    if len(workbook.worksheets) > 1:
        supplement_sheet = workbook.worksheets[1]
        header_seen = False
        for source_row, cells in enumerate(supplement_sheet.iter_rows(values_only=True), start=1):
            row = [clean_text(cell) for cell in cells]
            padded = row + [""] * (5 - len(row))
            if not any(padded):
                continue
            if padded[:5] == ["序号", "专业大类", "专业类", "专业代码", "专业名称"]:
                header_seen = True
                logs.append(f"college supplement header row {source_row}")
                continue
            if not header_seen:
                continue

            discipline = padded[1]
            category = padded[2]
            code = padded[3]
            name = padded[4]
            if not code or not name:
                continue

            major = {
                "major_id": stable_id("major", "专科", code, name),
                "major_code": code,
                "major_name": name,
                "major_category": category,
                "major_discipline": discipline,
                "degree_level": "专科",
                "degree": "",
                "duration": "",
                "subject_requirement": "",
                "description": "",
                "source_file": COLLEGE_MAJOR_SOURCE.name,
                "source_row": source_row,
            }
            if major["major_id"] not in seen_ids:
                majors.append(major)
                seen_ids.add(major["major_id"])

    logs.append(f"college majors parsed: {len(majors)}")
    return majors


def convert() -> ConversionResult:
    logs: list[str] = []
    schools = convert_schools(logs)
    majors = convert_undergrad_majors(logs) + convert_college_majors(logs)
    return ConversionResult(schools=schools, majors=majors, logs=logs)


def main() -> None:
    result = convert()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    schools_path = OUTPUT_DIR / "schools.json"
    majors_path = OUTPUT_DIR / "majors.json"

    write_json(schools_path, result.schools)
    write_json(majors_path, result.majors)

    print("Conversion completed.")
    print(f"schools.json: {schools_path}")
    print(f"majors.json: {majors_path}")
    print(f"schools count: {len(result.schools)}")
    print(f"majors count: {len(result.majors)}")
    print("Logs:")
    for line in result.logs[:50]:
        print(f"- {line}")
    if len(result.logs) > 50:
        print(f"- ... {len(result.logs) - 50} more log lines")


if __name__ == "__main__":
    main()
