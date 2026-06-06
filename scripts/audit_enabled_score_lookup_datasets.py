from __future__ import annotations

import json
import random
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pdfplumber

from national_extraction_common import ROOT, dump_json, iso_now, load_json, normalize_text, safe_float, safe_int, write_markdown

APP_INDEX = ROOT / "app" / "public" / "data" / "score-lookup" / "index.json"
AUDIT_JSON = ROOT / "data_work" / "enabled_score_lookup_audit.json"
AUDIT_MD = ROOT / "docs" / "enabled_score_lookup_audit_report.md"
ADMISSIONS_FOLDERS = [
    ROOT / "data_work" / "national_admissions_normalized",
    ROOT / "data_work" / "remaining_admissions_normalized",
]
RANK_FOLDERS = [ROOT / "data_work" / "national_rank_tables_normalized"]
WATERMARK_TOKEN_RE = re.compile(r"(^|\s)[江西省教育考试院](?=\s|$)")


@dataclass
class DatasetBundle:
    dataset: dict[str, Any]
    admissions_path: Path
    admissions_records: list[dict[str, Any]]
    rank_path: Path | None
    rank_records: list[dict[str, Any]]


def clean_pdf_text(value: Any) -> str:
    text = normalize_text(value).replace("\n", " ")
    text = WATERMARK_TOKEN_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def digits_only(value: Any) -> str:
    return "".join(ch for ch in clean_pdf_text(value) if ch.isdigit())


def find_dataset_file(folders: list[Path], province: str, year: int, subject_type: str, batch: str | None = None) -> Path:
    for folder in folders:
        for path in folder.glob("*.json"):
            data = load_json(path)
            if not isinstance(data, list) or not data:
                continue
            first = data[0]
            if (
                normalize_text(first.get("province")) == province
                and safe_int(first.get("year")) == year
                and normalize_text(first.get("subject_type")) == subject_type
                and (batch is None or normalize_text(first.get("batch")) == batch)
            ):
                return path
    raise FileNotFoundError(f"dataset source not found: {province}/{year}/{subject_type}/{batch or '-'}")


def load_enabled_datasets() -> list[dict[str, Any]]:
    payload = load_json(APP_INDEX)
    return [item for item in payload["datasets"] if item.get("enabled")]


def build_enabled_bundles() -> list[DatasetBundle]:
    bundles: list[DatasetBundle] = []
    for dataset in load_enabled_datasets():
        admissions_path = find_dataset_file(
            ADMISSIONS_FOLDERS,
            dataset["province"],
            int(dataset["year"]),
            dataset["subject_type"],
            dataset["batch"],
        )
        rank_path = None
        rank_records: list[dict[str, Any]] = []
        if dataset.get("rank_table_file"):
            rank_path = find_dataset_file(
                RANK_FOLDERS,
                dataset["province"],
                int(dataset["year"]),
                dataset["subject_type"],
                None,
            )
            rank_records = load_json(rank_path)
        bundles.append(
            DatasetBundle(
                dataset=dataset,
                admissions_path=admissions_path,
                admissions_records=load_json(admissions_path),
                rank_path=rank_path,
                rank_records=rank_records,
            )
        )
    return bundles


def pick_sample_indexes(records: list[dict[str, Any]], size: int = 30) -> list[int]:
    indexes = list(range(len(records)))
    rng = random.Random(20260606)
    sampled = set(rng.sample(indexes, min(size, len(indexes))))

    with_score = [idx for idx, record in enumerate(records) if safe_float(record.get("min_score")) is not None]
    with_score_sorted = sorted(with_score, key=lambda idx: float(records[idx]["min_score"]))
    sampled.update(with_score_sorted[:10])
    sampled.update(with_score_sorted[-10:])

    anomaly_indexes = []
    for idx, record in enumerate(records):
        school_name = normalize_text(record.get("school_name"))
        subject_type = normalize_text(record.get("subject_type"))
        min_rank = safe_int(record.get("min_rank"))
        min_score = safe_float(record.get("min_score"))
        if (
            not school_name
            or school_name == subject_type
            or school_name.endswith(subject_type)
            or min_rank is None
            or (min_score is not None and min_score <= 200)
        ):
            anomaly_indexes.append(idx)
    sampled.update(anomaly_indexes[:20])
    return sorted(sampled)


def build_excel_row_cache() -> dict[tuple[str, str, int], dict[str, Any]]:
    return {}


def get_excel_source_row(record: dict[str, Any], cache: dict[tuple[str, str, int], dict[str, Any]]) -> dict[str, Any]:
    import openpyxl

    source_file = Path(record["source_file"])
    source_sheet = normalize_text(record.get("source_sheet")) or "Sheet1"
    source_row = safe_int(record.get("source_row"))
    if source_row is None:
        raise ValueError("missing source_row")
    cache_key = (str(source_file), source_sheet, source_row)
    if cache_key in cache:
        return cache[cache_key]

    workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=True)
    worksheet = workbook[source_sheet if source_sheet in workbook.sheetnames else workbook.sheetnames[0]]
    headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    values = next(worksheet.iter_rows(min_row=source_row, max_row=source_row, values_only=True))
    row_map = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
    cache[cache_key] = row_map
    return row_map


def get_pdf_source_row(
    record: dict[str, Any],
    cache: dict[tuple[str, int], tuple[list[str], list[str]]],
) -> tuple[list[str], list[str]]:
    source_file = Path(record["source_file"])
    source_page = safe_int(record.get("source_page"))
    source_row = safe_int(record.get("source_row"))
    if source_page is None or source_row is None:
        raise ValueError("missing source_page/source_row")
    cache_key = (str(source_file), source_page)
    if cache_key not in cache:
        with pdfplumber.open(source_file) as pdf:
            page = pdf.pages[source_page - 1]
            tables = page.extract_tables()
        if not tables:
            raise ValueError("no table extracted from source pdf page")
        table = max(tables, key=len)
        headers = [clean_pdf_text(cell) for cell in table[0]]
        cache[cache_key] = (headers, table)

    headers, table = cache[cache_key]
    if source_row - 1 >= len(table):
        raise IndexError("source_row exceeds extracted table rows")
    return headers, table[source_row - 1]


def corrected_admissions_from_source(
    record: dict[str, Any],
    excel_cache: dict[tuple[str, str, int], dict[str, Any]],
    pdf_cache: dict[tuple[str, int], tuple[list[str], list[str]]],
) -> dict[str, Any]:
    source_file = normalize_text(record.get("source_file")).lower()
    batch = normalize_text(record.get("batch"))
    subject_type = normalize_text(record.get("subject_type"))
    year = safe_int(record.get("year"))

    if source_file.endswith(".xlsx") or source_file.endswith(".xlsm"):
        row = get_excel_source_row(record, excel_cache)
        return {
            "school_name": normalize_text(row.get("学校")),
            "major_group_code": "",
            "major_group_name": "",
            "major_name": normalize_text(row.get("专业")),
            "min_score": safe_float(row.get("最低分")),
            "min_rank": safe_int(row.get("最低位次")),
            "batch": normalize_text(row.get("批次")) or batch,
            "subject_type": normalize_text(row.get("科类")) or subject_type,
            "year": safe_int(row.get("年份")) or year,
        }

    if source_file.endswith(".pdf"):
        headers, row = get_pdf_source_row(record, pdf_cache)
        row_map = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        return {
            "school_name": clean_pdf_text(row_map.get("院校名称")),
            "major_group_code": digits_only(row_map.get("专业组 代号")) or digits_only(row_map.get("专业组代号")),
            "major_group_name": clean_pdf_text(row_map.get("专业组 名称")) or clean_pdf_text(row_map.get("专业组名称")),
            "major_name": "",
            "min_score": safe_float(clean_pdf_text(row_map.get("投档线"))),
            "min_rank": safe_int(clean_pdf_text(row_map.get("最低 投档排名")) or clean_pdf_text(row_map.get("最低投档排名"))),
            "batch": batch,
            "subject_type": clean_pdf_text(row_map.get("科类")) or subject_type,
            "year": year,
        }

    raise ValueError(f"unsupported source file for audit: {source_file}")


def audit_admissions_bundle(bundle: DatasetBundle) -> dict[str, Any]:
    sample_indexes = pick_sample_indexes(bundle.admissions_records)
    excel_cache: dict[tuple[str, str, int], dict[str, Any]] = {}
    pdf_cache: dict[tuple[str, int], tuple[list[str], list[str]]] = {}

    passed = 0
    failed = 0
    sample_notes = []
    recoverable_mismatch_count = 0

    for idx in sample_indexes:
        record = bundle.admissions_records[idx]
        try:
            corrected = corrected_admissions_from_source(record, excel_cache, pdf_cache)
            checks = {
                "school_name_ok": bool(corrected["school_name"]),
                "batch_ok": corrected["batch"] == normalize_text(bundle.dataset["batch"]),
                "subject_type_ok": corrected["subject_type"] == normalize_text(bundle.dataset["subject_type"]),
                "year_ok": corrected["year"] == int(bundle.dataset["year"]),
                "min_score_ok": corrected["min_score"] is not None,
            }
            row_passed = all(checks.values())
            current_school = normalize_text(record.get("school_name"))
            current_score = safe_float(record.get("min_score"))
            current_rank = safe_int(record.get("min_rank"))
            if (
                current_school != corrected["school_name"]
                or current_score != corrected["min_score"]
                or current_rank != corrected["min_rank"]
            ):
                recoverable_mismatch_count += 1
            if row_passed:
                passed += 1
            else:
                failed += 1
            if len(sample_notes) < 12:
                sample_notes.append(
                    {
                        "source_row": record.get("source_row"),
                        "source_page": record.get("source_page"),
                        "current_school_name": current_school,
                        "corrected_school_name": corrected["school_name"],
                        "current_min_score": current_score,
                        "corrected_min_score": corrected["min_score"],
                        "current_min_rank": current_rank,
                        "corrected_min_rank": corrected["min_rank"],
                        "passed": row_passed,
                    }
                )
        except Exception as exc:  # pragma: no cover - defensive
            failed += 1
            if len(sample_notes) < 12:
                sample_notes.append(
                    {
                        "source_row": record.get("source_row"),
                        "source_page": record.get("source_page"),
                        "error": str(exc),
                        "passed": False,
                    }
                )

    major_issues = []
    minor_issues = []
    if recoverable_mismatch_count:
        major_issues.append(
            f"抽检样本中有 {recoverable_mismatch_count} 条当前前端候选记录与原始 source row 不一致，发布包必须按 source row 重建。"
        )
    if failed:
        major_issues.append(f"仍有 {failed} 条样本无法稳定回查或关键字段不完整。")
    missing_rank_count = sum(1 for idx in sample_indexes if safe_int(bundle.admissions_records[idx].get("min_rank")) is None)
    if missing_rank_count:
        minor_issues.append(f"抽检样本中有 {missing_rank_count} 条记录缺少 min_rank，上线后只能进入“仅分数参考”。")

    release_decision = "keep_enabled" if passed / max(len(sample_indexes), 1) >= 0.8 else "needs_more_review"
    return {
        "dataset_id": bundle.dataset["dataset_id"],
        "sample_count": len(sample_indexes),
        "checked_count": len(sample_indexes),
        "passed_count": passed,
        "failed_count": failed,
        "major_issues": major_issues,
        "minor_issues": minor_issues,
        "release_decision": release_decision,
        "sample_notes": sample_notes,
    }


def audit_rank_bundle(bundle: DatasetBundle) -> dict[str, Any]:
    if not bundle.rank_records:
        return {
            "records": 0,
            "score_complete": False,
            "rank_complete": False,
            "monotonic": False,
            "sampled_points": [],
        }

    valid_records = [
        item
        for item in bundle.rank_records
        if safe_float(item.get("score")) is not None and safe_int(item.get("rank")) is not None
    ]
    score_complete = len(valid_records) / max(len(bundle.rank_records), 1) >= 0.95
    rank_complete = len(valid_records) / max(len(bundle.rank_records), 1) >= 0.95

    monotonic = True
    for prev, curr in zip(valid_records, valid_records[1:]):
        prev_score = safe_float(prev.get("score"))
        curr_score = safe_float(curr.get("score"))
        prev_rank = safe_int(prev.get("rank"))
        curr_rank = safe_int(curr.get("rank"))
        if None in {prev_score, curr_score, prev_rank, curr_rank}:
            monotonic = False
            break
        if curr_score < prev_score and curr_rank < prev_rank:
            monotonic = False
            break

    sample_pool = valid_records if valid_records else bundle.rank_records
    sample_indexes = sorted(
        {
            0,
            1,
            2,
            max(len(sample_pool) // 4, 0),
            max(len(sample_pool) // 2, 0),
            max((len(sample_pool) * 3) // 4, 0),
            max(len(sample_pool) - 3, 0),
            max(len(sample_pool) - 2, 0),
            max(len(sample_pool) - 1, 0),
            min(10, len(sample_pool) - 1),
        }
    )
    sampled_points = [
        {
            "score": sample_pool[idx]["score"],
            "rank": sample_pool[idx]["rank"],
            "source_page": sample_pool[idx].get("source_page"),
            "source_row": sample_pool[idx].get("source_row"),
        }
        for idx in sample_indexes
        if 0 <= idx < len(sample_pool)
    ]
    return {
        "records": len(bundle.rank_records),
        "score_complete": score_complete,
        "rank_complete": rank_complete,
        "monotonic": monotonic,
        "sampled_points": sampled_points,
    }


def main() -> None:
    bundles = build_enabled_bundles()
    dataset_audits = []
    rank_audits = {}
    release_keep_ids = []

    for bundle in bundles:
        admissions_audit = audit_admissions_bundle(bundle)
        rank_audit = audit_rank_bundle(bundle)
        rank_audits[bundle.dataset["dataset_id"]] = rank_audit

        if not (rank_audit["score_complete"] and rank_audit["rank_complete"] and rank_audit["monotonic"]):
            admissions_audit["major_issues"].append("对应 rank_table 未通过完整性或单调性检查。")
            admissions_audit["release_decision"] = "needs_more_review"

        if admissions_audit["release_decision"] == "keep_enabled":
            release_keep_ids.append(bundle.dataset["dataset_id"])

        dataset_audits.append(admissions_audit)

    payload = {
        "audited_at": iso_now(),
        "datasets": dataset_audits,
        "rank_audits": rank_audits,
        "release_keep_dataset_ids": release_keep_ids,
    }
    dump_json(AUDIT_JSON, payload)

    lines = [
        "# Enabled Score Lookup Audit Report",
        "",
        f"- 审核数据集数量：{len(dataset_audits)}",
        f"- 建议继续 enabled：{len(release_keep_ids)}",
        "",
    ]
    for item in dataset_audits:
        lines.extend(
            [
                f"## {item['dataset_id']}",
                "",
                f"- sample_count: {item['sample_count']}",
                f"- checked_count: {item['checked_count']}",
                f"- passed_count: {item['passed_count']}",
                f"- failed_count: {item['failed_count']}",
                f"- release_decision: {item['release_decision']}",
                f"- major_issues: {'；'.join(item['major_issues']) if item['major_issues'] else '无'}",
                f"- minor_issues: {'；'.join(item['minor_issues']) if item['minor_issues'] else '无'}",
                "",
                "### Sample Notes",
                "",
            ]
        )
        for sample in item["sample_notes"]:
            lines.append(f"- {json.dumps(sample, ensure_ascii=False)}")
        rank_item = rank_audits[item["dataset_id"]]
        lines.extend(
            [
                "",
                "### Rank Table Audit",
                "",
                f"- records: {rank_item['records']}",
                f"- score_complete: {rank_item['score_complete']}",
                f"- rank_complete: {rank_item['rank_complete']}",
                f"- monotonic: {rank_item['monotonic']}",
                "",
            ]
        )
        for point in rank_item["sampled_points"][:10]:
            lines.append(f"- sampled_point: {json.dumps(point, ensure_ascii=False)}")
        lines.append("")

    write_markdown(AUDIT_MD, "\n".join(lines))
    print(f"audited {len(dataset_audits)} enabled datasets; keep_enabled={len(release_keep_ids)}")


if __name__ == "__main__":
    main()
